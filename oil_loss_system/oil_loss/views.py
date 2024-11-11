from django.shortcuts import render, redirect
from .models import OilProduct, Storage, Transaction
from django.http import HttpResponse

# Просмотр нефтепродуктов
def oil_product_list(request):
    products = OilProduct.objects.all()
    return render(request, 'oil_loss/oil_product_list.html', {'products': products})

# Добавление нефтепродукта
def add_oil_product(request):
    if request.method == 'POST':
        name = request.POST['name']
        density = float(request.POST['density'])
        evaporation_rate = float(request.POST['evaporation_rate'])
        OilProduct.objects.create(name=name, density=density, evaporation_rate=evaporation_rate)
        return redirect('oil_product_list')
    return render(request, 'oil_loss/add_oil_product.html')

# Список транзакций
def transaction_list(request):
    transactions = Transaction.objects.all()
    total_loss = sum([t.calculate_loss() for t in transactions])
    return render(request, 'oil_loss/transaction_list.html', {'transactions': transactions, 'total_loss': total_loss})


import io
from django.http import FileResponse
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from .models import Transaction

# Генерация PDF отчета
def generate_pdf_report(request):
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer)

    transactions = Transaction.objects.all()
    y = 800
    for transaction in transactions:
        p.drawString(100, y, f"Дата: {transaction.date} | Нефтепродукт: {transaction.oil_product.name} | Начальный объем: {transaction.initial_volume} м³ | Естественная убыль: {transaction.natural_loss:.2f} м³")
        y -= 20

    p.showPage()
    p.save()
    buffer.seek(0)

    return FileResponse(buffer, as_attachment=True, filename='report.pdf')

# Генерация Excel отчета
def generate_excel_report(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Отчет по транзакциям'

    headers = ['Дата', 'Нефтепродукт', 'Начальный объем', 'Конечный объем', 'Естественная убыль']
    worksheet.append(headers)

    transactions = Transaction.objects.all()
    for transaction in transactions:
        row = [
            transaction.date,
            transaction.oil_product.name,
            transaction.initial_volume,
            transaction.final_volume,
            transaction.natural_loss
        ]
        worksheet.append(row)

    response = FileResponse(io.BytesIO(), as_attachment=True, filename='report.xlsx')
    workbook.save(response.stream)
    return response

from django.urls import path
from . import views

urlpatterns = [
    path('report/pdf/', views.generate_pdf_report, name='generate_pdf_report'),
    path('report/excel/', views.generate_excel_report, name='generate_excel_report'),
]


from rest_framework import viewsets
from .models import OilProduct, Storage, Transaction
from .serializers import OilProductSerializer, StorageSerializer, TransactionSerializer

class OilProductViewSet(viewsets.ModelViewSet):
    queryset = OilProduct.objects.all()
    serializer_class = OilProductSerializer

class StorageViewSet(viewsets.ModelViewSet):
    queryset = Storage.objects.all()
    serializer_class = StorageSerializer

class TransactionViewSet(viewsets.ModelViewSet):
    queryset = Transaction.objects.all()
    serializer_class = TransactionSerializer


from django.shortcuts import render
from .forms import TransactionForm


def add_transaction(request):
    if request.method == 'POST':
        form = TransactionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('success')  # Направить на другую страницу после успешного добавления
    else:
        form = TransactionForm()

    return render(request, 'add_transaction.html', {'form': form})


from rest_framework import viewsets
from .models import Transaction
from .serializers import TransactionSerializer

class TransactionViewSet(viewsets.ModelViewSet):
    queryset = Transaction.objects.all()
    serializer_class = TransactionSerializer


from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth import login, logout
from django.shortcuts import render, redirect

def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')
    else:
        form = UserCreationForm()
    return render(request, 'register.html', {'form': form})

def user_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('home')
    else:
        form = AuthenticationForm()
    return render(request, 'login.html', {'form': form})

def user_logout(request):
    logout(request)
    return redirect('login')

from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

def generate_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="oil_loss_report.pdf"'

    c = canvas.Canvas(response, pagesize=letter)
    c.drawString(100, 750, "Отчет по естественной убыли нефти")
    c.drawString(100, 730, "Дата: 2024-11-11")
    c.drawString(100, 710, "Показатели убыли нефти и нефтепродуктов:")

    # Пример данных
    c.drawString(100, 690, "Объем начальный: 1000 м³")
    c.drawString(100, 670, "Объем конечный: 950 м³")

    c.showPage()
    c.save()
    return response


from openpyxl import Workbook
from django.http import HttpResponse

def generate_excel(request):
    # Создаем книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"

    # Заголовки
    ws['A1'] = 'Наименование'
    ws['B1'] = 'Начальный объем (м³)'
    ws['C1'] = 'Конечный объем (м³)'

    # Пример данных
    ws['A2'] = 'Нефтепродукт 1'
    ws['B2'] = 1000
    ws['C2'] = 950

    # Ответ для скачивания Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="oil_loss_report.xlsx"'
    wb.save(response)
    return response


import matplotlib.pyplot as plt
from django.http import HttpResponse

def generate_chart(request):
    # Пример данных
    labels = ['Изначальный объем', 'Конечный объем']
    values = [1000, 950]

    # Генерация графика
    plt.bar(labels, values)
    plt.title('Потери нефти')

    # Сохранение графика в HTTP-ответ
    response = HttpResponse(content_type='image/png')
    plt.savefig(response, format="png")
    return response


from rest_framework import viewsets
from .models import Transaction
from .serializers import TransactionSerializer

class TransactionViewSet(viewsets.ModelViewSet):
    queryset = Transaction.objects.all()
    serializer_class = TransactionSerializer


from django.shortcuts import render

def home(request):
    return render(request, 'home.html')

from django.shortcuts import render
from .models import Transaction

def transaction_list(request):
    query = request.GET.get('q', '')  # Получаем параметр поиска из URL
    transactions = Transaction.objects.all()  # Получаем все транзакции

    if query:
        transactions = transactions.filter(oil_product__name__icontains=query)  # Фильтруем по названию нефтепродукта

    return render(request, 'transactions.html', {'transactions': transactions, 'query': query})


from django.shortcuts import render

def profile(request):
    return render(request, 'profile.html')

from django.shortcuts import render

# Представление для base.html
def show_base(request):
    return render(request, 'base.html')
